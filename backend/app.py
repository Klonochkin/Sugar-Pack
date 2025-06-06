import hashlib
import os
import zipfile
from fastapi import FastAPI, Form, HTTPException, Request, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.styles import Alignment

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

client = MongoClient("db", 27017)
# client.drop_database('typeProduct')
# client.drop_database('product')
# client.drop_database('feature')
# client.drop_database('user')
# client.drop_database('session')
# client.drop_database('feedback')
# client.drop_database('order')
# client.drop_database('material')
dbTypeProduct = client.typeProduct
postTypeProduct = dbTypeProduct.posts

dbProduct = client.product
postProduct = dbProduct.posts

dbFeature = client.feature
postFeature = dbFeature.posts

dbUser = client.user
postUser = dbUser.posts

dbSession = client.session
postSession = dbSession.posts

dbFeedback = client.feedback
postFeedback = dbFeedback.posts

dbOrder = client.order
postOrder = dbOrder.posts

dbMaterial = client.material
postMaterial = dbMaterial.posts

@app.get('/')
async def welcome(request: Request):
    # dataSend = postUser.find({})
    # posts_list = []
    # for post in dataSend:
    #     post_dict = dict(post)
    #     del post_dict['_id']
    #     posts_list.append(post_dict)
    # print(posts_list)
    cookies = request.cookies
    session_value = cookies.get("session")
    res = postSession.find_one({"session": session_value})
    if(res==None):
        return [{"role" : "none"}]
    else:
        dataSend = postUser.find({"id": res["id"]})
        posts_list = []
        for post in dataSend:
            post_dict = dict(post)
            del post_dict['_id']
            del post_dict['password']
            del post_dict['salt']
            posts_list.append(post_dict)
        return posts_list

@app.get('/api/get-type-catalog')
async def getCatalog():
    dataSend = postTypeProduct.find({})
    posts_list = []
    for post in dataSend:
        post_dict = dict(post)
        del post_dict['_id']
        posts_list.append(post_dict)
    return posts_list

@app.get('/api/get-product-catalog/{id}')
async def getFullCatalog(id: str):
    dataSend = postProduct.find({'idTypeProduct': id})
    posts_list = []
    for post in dataSend:
        post_dict = dict(post)
        del post_dict['_id']
        posts_list.append(post_dict)
    return posts_list

@app.get('/api/get-product/{id}')
async def getProduct(id: str):
    dataSend = postProduct.find({'id': id})
    posts_list = []
    for post in dataSend:
        post_dict = dict(post)
        del post_dict['_id']
        posts_list.append(post_dict)
    return posts_list

@app.get('/api/get-feature/{id}')
async def getProduct(id: str):
    dataSend = postFeature.find({'idProduct': id})
    posts_list = []
    for post in dataSend:
        post_dict = dict(post)
        del post_dict['_id']
        posts_list.append(post_dict)
    return posts_list

@app.post('/api/send-feedback')
async def sendFeedback(request: Request):
    data = await request.json()
    id = data["id"]
    name = data["name"]
    email = data["email"]
    phone = data["phone"]
    message = data["message"]
    n=postFeedback.count_documents({})
    userId = n+1
    if(postFeedback.count_documents({}) == n):
        post = {
            "id": userId,
            "productId": int(id),
            "name": name,
            "email": email,
            "phone":phone,
            "message": message,
        }
        postFeedback.insert_one(post)
    return {"message": "Отзыв отправлен"}

@app.get('/api/get-feedback/{id}')
async def getFeedback(id: str):
    dataSend = postFeedback.find({"productId": int(id)})
    posts_list = []
    for post in dataSend:
        post_dict = dict(post)
        del post_dict['_id']
        del post_dict['email']
        del post_dict['phone']
        posts_list.append(post_dict)
    return posts_list

@app.get('/api/get-feedback/')
async def getFeedback():
    dataSend = postFeedback.find({})
    posts_list = []
    for post in dataSend:
        post_dict = dict(post)
        del post_dict['_id']
        posts_list.append(post_dict)
    return posts_list


@app.post("/api/signIn")
async def signIn(request: Request):
    data = await request.json()
    login = data["login"]
    password = data["password"]
    res = postUser.find_one({'login': login})
    if(res==None):
        res = postUser.find_one({'email': login})
        if(res == None):
            raise HTTPException(status_code=401, detail="Аккаунт не найден")

    salt = res['salt']
    truePassword = res['password']
    checkPassword = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt, 100000).hex()

    if(truePassword != checkPassword):
        raise HTTPException(status_code=422, detail="Не правильный пароль")

    hashSession = hashlib.pbkdf2_hmac('sha256', login.encode('utf-8'),os.urandom(16).hex().encode('utf-8'), 100000).hex()
    content = {"message": "true"}
    response = JSONResponse(content=content)
    response.set_cookie(key="session", value=hashSession)
    filterDelete = {'id':res["id"]}
    postSession.delete_one(filterDelete)
    n=postSession.count_documents({})

    if(postSession.count_documents({}) == n):
        post = {
            "id": res["id"],
            "session": hashSession
        }
        res = postSession.insert_one(post)

    return response

@app.post("/api/signUp")
async def signUp(request: Request):
    data = await request.json()
    login = data["login"]
    email = data["email"]
    phone = data["phone"]
    password = data["password"]
    res = postUser.find_one({'email': email})
    res1 = postUser.find_one({'login': login})
    if(res!=None):
        raise HTTPException(status_code=401, detail="Аккаунт уже существует")
    else:
        if(res1 != None):
            raise HTTPException(status_code=401, detail="Аккаунт уже существует")
    if(len(password)<8):
        raise HTTPException(status_code=422, detail="Пароль слишком короткий")
    n=postUser.count_documents({})
    salt = os.urandom(16)
    hashPassword = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt.hex().encode('utf-8'), 100000)
    userId = n+1
    if( postUser.count_documents({}) == n):
        post = {
            "id": userId,
            "login" : login,
            "email": email,
            "phone": phone,
            "role": "user",
            "password":hashPassword.hex(),
            "salt":salt.hex().encode('utf-8')
        }
        postUser.insert_one(post)
    return {"message": "Регистрация успешна"}

@app.post("/api/signOut")
async def signOut(request: Request):
    content = {"message": "true"}
    session = request.cookies.get("session")
    response = JSONResponse(content=content)
    filterDelete = {'session':session}
    postSession.delete_one(filterDelete)
    return response

@app.post('/api/send-order')
async def sendOrder(request: Request):
    cookies = request.cookies
    session_value = cookies.get("session")
    res = postSession.find_one({"session": session_value})
    if(res==None):
        raise HTTPException(status_code=403, detail="Пользователь не вошёл")
    data = await request.json()
    name = data["name"]
    phone = data["phone"]
    quantity = data["quantity"]
    n=postOrder.count_documents({})
    userId = res["id"]
    orderId = n+1
    if( postOrder.count_documents({}) == n):
        post = {
            "id": orderId,
            "idUser":userId,
            "name": name,
            "phone": phone,
            "quantity":quantity,
            "status": "processing",
        }
        postOrder.insert_one(post)
    return {"message": "Заказ добавлен"}

@app.get('/api/get-order')
async def getOrder(request: Request):
    cookies = request.cookies
    session_value = cookies.get("session")
    res = postSession.find_one({"session": session_value})
    if(res==None):
        return [{'message': 'noob'}]
    else:
        userInfo = postUser.find_one({"id":res["id"]})
        if(userInfo["role"] == "admin"):
            dataSend = postOrder.find({})
            posts_list = []
            for post in dataSend:
                post_dict = dict(post)
                del post_dict['_id']
                posts_list.append(post_dict)
            return posts_list
        else:
            dataSend = postOrder.find({"idUser":res["id"]})
            posts_list = []
            for post in dataSend:
                post_dict = dict(post)
                del post_dict['_id']
                posts_list.append(post_dict)
            return posts_list

@app.post('/api/update-order')
async def updateOrder(request: Request):
    data = await request.json()
    id = data["id"]
    status = data["status"]
    cookies = request.cookies
    session_value = cookies.get("session")
    res = postSession.find_one({"session": session_value})
    if(res==None):
        return [{'message': 'noob'}]
    else:
        userInfo = postUser.find_one({"id":res["id"]})
        if(userInfo["role"] == "admin"):
            filter = {'id':int(id)}
            postOrder.update_many(filter, {'$set': {'status': status}})
            return [{'message': 'ok'}]
        else:
            if(userInfo["role"] == "user" and status=="cansel"):
                checkIsOwner = postOrder.find_one({"id":id})
                if(checkIsOwner["idUser"] == userInfo["id"]):
                    filter = {'id':int(id)}
                    postOrder.update_many(filter, {'$set': {'status': status}})
                    return [{'message': 'ok'}]
    return [{'message': '403 erroe'}]

@app.post('/api/send-material/')
async def sendMaterial(request: Request):
    cookies = request.cookies
    session_value = cookies.get("session")
    res = postSession.find_one({"session": session_value})
    if(res==None):
        return [{'message': '403 error'}]
    userInfo = postUser.find_one({"id":res["id"]})
    if(userInfo["role"] != "admin"):
        return [{'message': '411 error'}]
    form_data = await request.form()
    count = int(form_data.get("count"))
    orderId = form_data.get("orderId")
    for i in range(count):
        n=postMaterial.count_documents({})
        materialId = n+1
        print("if")
        if(postMaterial.count_documents({}) == n):
            post = {
                "id": materialId,
                "idOrder":orderId,
                "material": form_data.get("material" + str(i)),
                "quantity": form_data.get("quantity" + str(i)),
            }
            postMaterial.insert_one(post)
    return {"message": "true"}

@app.get('/api/get-material/')
async def getMaterial(request: Request):
    cookies = request.cookies
    session_value = cookies.get("session")
    res = postSession.find_one({"session": session_value})
    if(res==None):
        return [{'message': '403 error'}]
    userInfo = postUser.find_one({"id":res["id"]})
    if(userInfo["role"] != "admin"):
        return [{'message': '411 error'}]
    dataSend = postMaterial.find({})
    posts_list = []
    for post in dataSend:
        post_dict = dict(post)
        del post_dict['_id']
        posts_list.append(post_dict)
    return posts_list

@app.get("/api/export-report/")
async def export(request:Request):
    cookies = request.cookies
    session_value = cookies.get("session")
    res = postSession.find_one({"session": session_value})
    if(res==None):
        return [{'message': '403 error'}]
    userInfo = postUser.find_one({"id":res["id"]})
    if(userInfo["role"] != "admin"):
        return [{'message': '411 error'}]
    wb = Workbook()
    ws = wb.active

    res = postMaterial.find({}).sort("idOrder")

    orderId = []
    name = []
    count = []
    material = []
    quantity = []

    for document in res:
        orderInfo = postOrder.find_one({"id":int(document['idOrder'])})
        orderId.append(document['idOrder'])
        name.append(orderInfo["name"])
        count.append(orderInfo["quantity"])
        material.append(document['material'])
        quantity.append(document['quantity'])

    with zipfile.ZipFile('Material spent.zip', 'w') as zip_file:
        pass


    wb.save('Material spent.xlsx')

    ws['A1'] = '№ заказа'
    ws['B1'] = 'Товар'
    ws['C1'] = 'Количество товара, шт'
    ws['D1'] = 'Материал'
    ws['E1'] = 'Количество, шт'
    for i in range(len(orderId)):
        if(i == 0 or orderId[i] != orderId[i-1]):
            ws.cell(row=i+2, column=1).value = orderId[i]
            ws.cell(row=i+2, column=2).value = name[i]
            ws.cell(row=i+2, column=3).value = count[i]
            ws.cell(row=i+2, column=4).value = material[i]
            ws.cell(row=i+2, column=5).value = quantity[i]
        else:
            ws.cell(row=i+2, column=4).value = material[i]
            ws.cell(row=i+2, column=5).value = quantity[i]

    for col_idx, col in enumerate(ws.columns, start=1):
        max_width = 0
        for cell in col:
            if cell.value:
                width = len(str(cell.value)) + 2
                if width > max_width:
                    max_width = width
                cell.alignment = Alignment(horizontal='center', vertical='top')
        ws.column_dimensions[chr(64 + col_idx)].width = max_width

    wb.save('Material spent.xlsx')

    with zipfile.ZipFile('Material spent.zip', 'a') as zip_file:
        zip_file.write("Material spent.xlsx")

    try:
        return Response(
            content=open("Material spent.zip", "rb").read(),
            media_type="application/zip",
            headers={
                "Content-Disposition": "attachment; filename=Material spent.zip",
                "Cache-Control": "max-age=3600",
                "Accept-Encoding": "gzip, compress, br"
            }
        )
    finally:
        os.remove("Material spent.xlsx")
        os.remove("Material spent.zip")

